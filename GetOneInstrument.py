#  Notes
# To Use you need to create an excel file to store RandomRawData.xlsx
# Then use the code to get one instrument at a time and put the data into the excel file (Which must already exist and be closed)
# Its slow and will crash if the RandomRawData.xlsx file is open
# It also has major problems if you enter the wrong data - sometimes it hangs excel and corrupts RRD excel file
# If this happens recreate a new RRD.xlsx file (I will work on these things)



from ibapi.client import EClient
from ibapi.wrapper import EWrapper
from ibapi.client import Contract
from ibapi.common import *
import time
import openpyxl

class TestApp(EWrapper, EClient):
    def __init__(self):
        EClient.__init__(self, self)

    def error(self, reqId: TickerId, errorCode: int, errorString: str):
        super().error(reqId, errorCode, errorString)
        print("Error. Id: ", reqId, " Code: ", errorCode, " Msg: ", errorString)

    def historicalData(self, reqId: int, bar: BarData):
        print("HistoricalData. ", reqId, " Date:", bar.date, "Open:", bar.open,
              "High:", bar.high, "Low:", bar.low, "Close:", bar.close, "Volume:", bar.volume,
              "Count:", bar.barCount, "WAP:", bar.average)
        rowData = (bar.date, bar.open, bar.high, bar.low, bar.close, bar.volume, bar.barCount, bar.average)

        # append data to file (I have erased the sheet and
        # entered a new one so the data will always be in the same place)
        r = wb[instrument].max_row
        r += 1
        for i, header in enumerate(rowData):
            wb[instrument].cell(row=r, column=i+1).value = header

    def historicalDataEnd(self, reqId: int, start: str, end: str):
        # finish getting data
        self.done = True


def main():

    app = TestApp()
    app.connect("127.0.0.1", 7497, 2)  # socket port is set in TWS or IB Gateway settings

    time.sleep(1)  # short sleep to allow connection

    contract = Contract()
    contract.symbol = instrument
    contract.secType = securitytype
    contract.exchange = exchange
    contract.currency = "USD"
    contract.lastTradeDateOrContractMonth = "201903"

    app.reqHistoricalData(1, contract, "", length, barSize, "TRADES", 1, 1, False, [])
    print(instrument)
    app.run()
    wb.save('RandomRawData.xlsx')

#global program
instrument = input("Enter the Instrument (AAPL... ) >")
securitytype = input("Enter the type of security (STK, FUT ...) >")
exchange = input("Enter the exchange (SMART... )  >")
length = input("Enter the length for the Data (1 D, 1800, S...) >")
barSize = input("Enter the bar size for the Data (30 mins, 1 day...) >")
wb = openpyxl.load_workbook('RandomRawData.xlsx')
all_sheets = wb.sheetnames

if instrument in all_sheets:
    wb.remove(wb[instrument])
    ws = wb.create_sheet(instrument)
    print(f"yes {instrument} is here")
else:
    print(f"no {instrument} is not here")
    ws = wb.create_sheet(instrument)


if __name__ == "__main__":
    main()


